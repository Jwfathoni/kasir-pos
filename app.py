from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
from starlette.middleware.sessions import SessionMiddleware
from datetime import datetime, date, timedelta
from calendar import monthrange
import pandas as pd
import io
import os
import shutil
from fastapi.responses import JSONResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

from db import engine, Base, get_db
from models import User, Product, Transaction, TransactionItem, Setting, StockUpdate
from auth import verify_password, require_login, hash_password
from sqlalchemy import text

Base.metadata.create_all(bind=engine)

# Auto-migration: Tambahkan kolom cost_price jika belum ada dan buat tabel stock_updates
def run_migrations():
    try:
        with engine.begin() as conn:  # Gunakan begin() untuk auto-commit
            # Cek dan tambahkan cost_price ke products
            try:
                result = conn.execute(text("PRAGMA table_info(products)"))
                columns = [row[1] for row in result]
                if 'cost_price' not in columns:
                    conn.execute(text("ALTER TABLE products ADD COLUMN cost_price FLOAT NOT NULL DEFAULT 0"))
                    print("[OK] Migration: Kolom cost_price ditambahkan ke tabel products")
            except Exception as e:
                print(f"Migration products: {e}")
            
            # Cek dan tambahkan cost_price ke transaction_items
            try:
                result = conn.execute(text("PRAGMA table_info(transaction_items)"))
                columns = [row[1] for row in result]
                if 'cost_price' not in columns:
                    conn.execute(text("ALTER TABLE transaction_items ADD COLUMN cost_price FLOAT NOT NULL DEFAULT 0"))
                    print("[OK] Migration: Kolom cost_price ditambahkan ke tabel transaction_items")
            except Exception as e:
                print(f"Migration transaction_items: {e}")
            
            # Cek dan buat tabel stock_updates jika belum ada
            try:
                result = conn.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='stock_updates'"))
                if not result.fetchone():
                    conn.execute(text("""
                        CREATE TABLE stock_updates (
                            id INTEGER NOT NULL PRIMARY KEY,
                            product_id INTEGER NOT NULL,
                            product_code VARCHAR NOT NULL,
                            product_name VARCHAR NOT NULL,
                            old_stock INTEGER NOT NULL DEFAULT 0,
                            new_stock INTEGER NOT NULL DEFAULT 0,
                            stock_added INTEGER NOT NULL DEFAULT 0,
                            cost_price FLOAT NOT NULL DEFAULT 0,
                            total_pengeluaran FLOAT NOT NULL DEFAULT 0,
                            created_at DATETIME,
                            updated_by VARCHAR
                        )
                    """))
                    print("[OK] Migration: Tabel stock_updates berhasil dibuat")
            except Exception as e:
                print(f"Migration stock_updates: {e}")
            
            # Cek dan tambahkan timezone ke settings
            try:
                result = conn.execute(text("PRAGMA table_info(settings)"))
                columns = [row[1] for row in result]
                if 'timezone' not in columns:
                    conn.execute(text("ALTER TABLE settings ADD COLUMN timezone VARCHAR DEFAULT 'WIB'"))
                    print("[OK] Migration: Kolom timezone ditambahkan ke tabel settings")
            except Exception as e:
                print(f"Migration settings timezone: {e}")
    except Exception as e:
        print(f"Error saat migration: {e}")

run_migrations()

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="change-this-secret")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

def format_idr(value):
    return "{:,}".format(int(value)).replace(",", ".")

templates.env.filters["format_idr"] = format_idr

# Timezone utility functions
TIMEZONE_MAP = {
    "WIB": "Asia/Jakarta",      # UTC+7
    "WITA": "Asia/Makassar",    # UTC+8
    "WIT": "Asia/Jayapura"      # UTC+9
}

def get_current_setting_timezone(db: Session) -> str:
    """Get timezone from settings, default to WIB"""
    settings = db.query(Setting).first()
    return settings.timezone if settings and settings.timezone else "WIB"

def get_timezone_offset(timezone_str: str) -> timedelta:
    """Get UTC offset untuk timezone Indonesia"""
    offsets = {"WIB": 7, "WITA": 8, "WIT": 9}
    hours = offsets.get(timezone_str, 7)
    return timedelta(hours=hours)

def get_current_time_with_tz(db: Session) -> datetime:
    """Get current time adjusted to setting timezone"""
    tz_name = get_current_setting_timezone(db)
    tz = pytz.timezone(TIMEZONE_MAP[tz_name])
    return datetime.now(tz)

def format_datetime_with_tz(dt: datetime, db: Session, format_str: str = "%d-%m-%Y %H:%M:%S") -> str:
    """Format datetime dengan timezone yang dipilih"""
    if dt is None:
        return "-"
    
    # Jika dt adalah naive datetime (tanpa timezone info), anggap sebagai UTC
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    
    tz_name = get_current_setting_timezone(db)
    tz = pytz.timezone(TIMEZONE_MAP[tz_name])
    dt_local = dt.astimezone(tz)
    return dt_local.strftime(format_str)

templates.env.filters["format_datetime_tz"] = format_datetime_with_tz

def make_trx_no(db: Session):
    today = get_current_time_with_tz(db).strftime("%Y%m%d")
    like = f"TRX-{today}-%"
    count = db.query(func.count(Transaction.id)).filter(Transaction.trx_no.like(like)).scalar() or 0
    return f"TRX-{today}-{count+1:04d}"

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=302)
    return RedirectResponse("/cashier", status_code=302)

# -------- LOGIN ----------
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    if not user or not verify_password(password, user.password_hash):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Username / password salah"})
    if user.display_name is None:
        user.display_name = user.username
        db.commit()
        db.refresh(user)
    request.session["user"] = {"username": user.username, "display_name": user.display_name, "role": user.role}
    return RedirectResponse("/cashier", status_code=302)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)

# -------- PRODUCTS ----------
@app.get("/products", response_class=HTMLResponse)
def products_page(request: Request, err: str = None, db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r
    products = db.query(Product).order_by(Product.id.desc()).all()
    error_msg = None
    if err == "code-exists":
        error_msg = "Kode produk sudah ada!"
    return templates.TemplateResponse("products.html", {"request": request, "products": products, "user": request.session["user"], "error": error_msg})

@app.post("/products/add")
def add_product(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    cost_price: float = Form(...),
    price: float = Form(...),
    stock: int = Form(0),
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r
    if db.query(Product).filter(Product.code == code).first():
        return RedirectResponse("/products?err=code-exists", status_code=302)
    db.add(Product(code=code, name=name, cost_price=cost_price, price=price, stock=stock, status="active"))
    db.commit()
    return RedirectResponse("/products", status_code=302)

@app.post("/products/update")
def update_product(
    request: Request,
    pid: int = Form(...),
    name: str = Form(...),
    cost_price: float = Form(...),
    price: float = Form(...),
    stock_add: int = Form(0),  # Jumlah stok yang ditambahkan (default 0)
    stock: int = Form(...),  # Stok saat ini (dari hidden input, untuk referensi)
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r
    p = db.query(Product).filter(Product.id == pid).first()
    if p:
        old_stock = p.stock
        
        # Update data produk
        p.name = name
        p.cost_price = cost_price
        p.price = price
        
        # Stok hanya bisa ditambah, tidak bisa dikurangi manual
        # Jika stock_add > 0, tambahkan ke stok lama
        if stock_add and stock_add > 0:
            new_stock = old_stock + stock_add
            p.stock = new_stock
            
            # Track stock update
            total_pengeluaran = cost_price * stock_add
            
            stock_update = StockUpdate(
                product_id=p.id,
                product_code=p.code,
                product_name=p.name,
                old_stock=old_stock,
                new_stock=new_stock,
                stock_added=stock_add,
                cost_price=cost_price,
                total_pengeluaran=total_pengeluaran,
                updated_by=request.session.get("user", {}).get("display_name", "System")
            )
            db.add(stock_update)
        # Jika stock_add = 0 atau tidak diisi, stok tetap sama (hanya update harga/nama)
        # Stok tidak bisa dikurangi manual, hanya berkurang saat transaksi
        
        db.commit()
    return RedirectResponse("/products", status_code=302)

@app.post("/products/update_name")
def update_product_name(
    request: Request,
    pid: int = Form(...),
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    """Endpoint khusus untuk update nama produk saja (inline edit)"""
    r = require_login(request)
    if r: return r
    p = db.query(Product).filter(Product.id == pid).first()
    if p:
        p.name = name.strip()
        if not p.name:
            return JSONResponse({"success": False, "message": "Nama produk tidak boleh kosong"}, status_code=400)
        db.commit()
        return JSONResponse({"success": True, "message": "Nama produk berhasil diupdate"})
    return JSONResponse({"success": False, "message": "Produk tidak ditemukan"}, status_code=404)

@app.post("/products/delete")
def delete_product(request: Request, pid: int = Form(...), db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r
    p = db.query(Product).filter(Product.id == pid).first()
    if p:
        db.delete(p)
        db.commit()
    return RedirectResponse("/products", status_code=302)

@app.post("/products/import_excel")
def import_products_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Hanya file Excel (.xlsx, .xls) yang diizinkan!")

    try:
        # Read the Excel file into a pandas DataFrame
        df = pd.read_excel(file.file, engine='openpyxl')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file Excel: {e}")

    required_columns = ['kode_produk', 'nama_produk', 'harga_asli', 'harga_jual', 'stok']
    if not all(col in df.columns for col in required_columns):
        raise HTTPException(status_code=400, detail=f"File Excel harus memiliki kolom: {', '.join(required_columns)}")

    products_added = 0
    products_updated = 0
    errors = []

    for index, row in df.iterrows():
        code = str(row['kode_produk']).strip()
        name = str(row['nama_produk']).strip()
        
        try:
            cost_price = float(row['harga_asli'])
            price = float(row['harga_jual'])
            stock = int(row['stok'])
        except ValueError:
            errors.append(f"Baris {index + 2}: Harga atau stok tidak valid untuk produk {code}.")
            continue

        if not code or not name or cost_price < 0 or price < 0 or stock < 0:
            errors.append(f"Baris {index + 2}: Data produk tidak lengkap atau tidak valid (kode, nama, harga_asli, harga_jual, stok).")
            continue

        existing_product = db.query(Product).filter(Product.code == code).first()

        if existing_product:
            existing_product.name = name
            existing_product.cost_price = cost_price
            existing_product.price = price
            existing_product.stock = stock
            products_updated += 1
        else:
            new_product = Product(code=code, name=name, cost_price=cost_price, price=price, stock=stock, status="active")
            db.add(new_product)
            products_added += 1
    
    db.commit()

    if errors:
        error_message = "Beberapa produk gagal diimpor:\n" + "\n".join(errors)
        return JSONResponse(status_code=400, content={
            "message": f"Berhasil mengimpor {products_added} produk baru dan memperbarui {products_updated} produk. Namun, ada beberapa kesalahan:",
            "detail": error_message
        })
    else:
        return JSONResponse(status_code=200, content={
            "message": f"Berhasil mengimpor {products_added} produk baru dan memperbarui {products_updated} produk."
        })

@app.get("/products/export_stock_template")
def export_stock_update_template(request: Request, db: Session = Depends(get_db)):
    """Export template Excel untuk update stok"""
    r = require_login(request)
    if r: return r
    
    products = db.query(Product).filter(Product.status == "active").order_by(Product.code).all()
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame([{
            "kode_produk": p.code,
            "nama_produk": p.name,
            "harga_asli": p.cost_price,
            "harga_jual": p.price,
            "stok_sekarang": p.stock,
            "stok_baru": p.stock,  # Default sama dengan stok sekarang
        } for p in products])
        df.to_excel(writer, sheet_name='Update Stok', index=False)
        
        # Apply styling
        workbook = writer.book
        ws = workbook['Update Stok']
        ws.freeze_panes = 'A2'
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Roboto")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Style header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border_style
                if cell.column_letter in ['C', 'D', 'E', 'F']:  # Kolom harga dan stok
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = right_align
        
        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > 0:
                ws.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    file_name = f"Template_Update_Stok_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename={file_name}"
    })

@app.post("/products/import_stock_update")
def import_stock_update_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import update stok dari Excel dan hitung total pengeluaran"""
    r = require_login(request)
    if r: return r

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Hanya file Excel (.xlsx, .xls) yang diizinkan!")

    try:
        df = pd.read_excel(file.file, engine='openpyxl')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file Excel: {e}")

    required_columns = ['kode_produk', 'harga_asli', 'harga_jual', 'stok_baru']
    if not all(col in df.columns for col in required_columns):
        raise HTTPException(status_code=400, detail=f"File Excel harus memiliki kolom: {', '.join(required_columns)}")

    products_updated = 0
    total_pengeluaran = 0
    errors = []
    stock_updates = []

    for index, row in df.iterrows():
        code = str(row['kode_produk']).strip()
        
        try:
            cost_price = float(row['harga_asli'])
            price = float(row['harga_jual'])
            new_stock = int(row['stok_baru'])
        except (ValueError, KeyError) as e:
            errors.append(f"Baris {index + 2}: Data tidak valid untuk produk {code}.")
            continue

        if not code or cost_price < 0 or price < 0 or new_stock < 0:
            errors.append(f"Baris {index + 2}: Data produk tidak lengkap atau tidak valid.")
            continue

        product = db.query(Product).filter(Product.code == code).first()
        if not product:
            errors.append(f"Baris {index + 2}: Produk dengan kode {code} tidak ditemukan.")
            continue

        old_stock = product.stock
        
        # Update produk
        product.cost_price = cost_price
        product.price = price
        product.stock = new_stock
        
        # Track stock update jika stok bertambah
        if new_stock > old_stock:
            stock_added = new_stock - old_stock
            pengeluaran = cost_price * stock_added
            total_pengeluaran += pengeluaran
            
            stock_update = StockUpdate(
                product_id=product.id,
                product_code=product.code,
                product_name=product.name,
                old_stock=old_stock,
                new_stock=new_stock,
                stock_added=stock_added,
                cost_price=cost_price,
                total_pengeluaran=pengeluaran,
                updated_by=request.session.get("user", {}).get("display_name", "System")
            )
            stock_updates.append(stock_update)
        
        products_updated += 1
    
    # Commit semua perubahan
    for su in stock_updates:
        db.add(su)
    db.commit()

    message = f"Berhasil mengupdate {products_updated} produk."
    if total_pengeluaran > 0:
        message += f" Total pengeluaran untuk update stok: Rp {total_pengeluaran:,.0f}".replace(",", ".")

    if errors:
        error_message = "Beberapa produk gagal diupdate:\n" + "\n".join(errors)
        return JSONResponse(status_code=400, content={
            "message": message,
            "detail": error_message,
            "total_pengeluaran": total_pengeluaran
        })
    else:
        return JSONResponse(status_code=200, content={
            "message": message,
            "total_pengeluaran": total_pengeluaran
        })


# -------- CASHIER ----------
@app.get("/api/timezone-info")
def get_timezone_info(db: Session = Depends(get_db)):
    """API untuk mendapatkan timezone dan waktu saat ini"""
    timezone_name = get_current_setting_timezone(db)
    current_time = get_current_time_with_tz(db)
    return {
        "timezone": timezone_name,
        "current_time": current_time.strftime("%d-%m-%Y %H:%M:%S"),
        "current_time_formatted": current_time.strftime("%A, %d %B %Y %H:%M:%S")
    }

@app.get("/cashier", response_class=HTMLResponse)
def cashier_page(request: Request, err: str = None, db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r
    products = db.query(Product).filter(Product.status == "active").order_by(Product.name.asc()).all()
    settings = db.query(Setting).first()
    error_msg = None
    if err == "empty":
        error_msg = "Keranjang kosong!"
    elif err == "paid-less":
        error_msg = "Jumlah bayar kurang dari total!"
    elif err == "invalid":
        error_msg = "Data keranjang tidak valid!"
    return templates.TemplateResponse("cashier.html", {"request": request, "products": products, "user": request.session["user"], "error": error_msg, "settings": settings})

@app.post("/checkout")
def checkout(
    request: Request,
    payment_method: str = Form(...),
    paid: float = Form(...),
    cart_json: str = Form(...),
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r

    import json
    try:
        cart = json.loads(cart_json)  # [{code,name,price,qty}]
    except (json.JSONDecodeError, ValueError):
        return RedirectResponse("/cashier?err=invalid", status_code=302)
    if not cart:
        return RedirectResponse("/cashier?err=empty", status_code=302)

    total = 0
    for it in cart:
        total += float(it["price"]) * int(it["qty"])

    change = paid - total
    if change < 0:
        return RedirectResponse("/cashier?err=paid-less", status_code=302)

    trx = Transaction(
        trx_no=make_trx_no(db),
        cashier=request.session["user"]["display_name"],
        payment_method=payment_method,
        total=total,
        paid=paid,
        change=change,
    )

    for it in cart:
        qty = int(it["qty"])
        price = float(it["price"])
        subtotal = qty * price
        # Kurangi stok kalau dipakai dan ambil cost_price
        p = db.query(Product).filter(Product.code == it["code"]).first()
        cost_price = p.cost_price if p else 0
        if p:
            p.stock = max(0, p.stock - qty)
        trx.items.append(TransactionItem(
            product_code=it["code"],
            product_name=it["name"],
            price=price,
            cost_price=cost_price,
            qty=qty,
            subtotal=subtotal
        ))

    db.add(trx)
    db.commit()
    return RedirectResponse(f"/receipt/{trx.id}", status_code=302)

# -------- RECEIPT ----------
@app.get("/receipt/{trx_id}", response_class=HTMLResponse)
def receipt_page(request: Request, trx_id: int, from_page: str = Query(None, alias="from"), mode: str = Query("daily"), db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r
    trx = db.query(Transaction).filter(Transaction.id == trx_id).first()
    settings = db.query(Setting).first()
    if not trx:
        return RedirectResponse("/cashier", status_code=302)
    # Ensure display_name is available in the session for the receipt page
    logged_in_user = db.query(User).filter(User.username == request.session["user"]["username"]).first()
    if logged_in_user and logged_in_user.display_name is None:
        logged_in_user.display_name = logged_in_user.username
        db.commit()
        db.refresh(logged_in_user)
        request.session["user"]["display_name"] = logged_in_user.display_name # Update session immediately
    
    # Tentukan URL kembali berdasarkan dari mana struk dibuka
    back_url = "/cashier"
    if from_page == "reports":
        back_url = f"/reports?mode={mode}"
    
    # Format created_at dengan timezone
    formatted_created_at = format_datetime_with_tz(trx.created_at, db, "%d/%m/%Y %H:%M")
    
    return templates.TemplateResponse("receipt.html", {
        "request": request, 
        "trx": trx, 
        "user": request.session["user"], 
        "settings": settings,
        "back_url": back_url,
        "formatted_created_at": formatted_created_at
    })

# -------- REPORTS ----------
@app.get("/reports", response_class=HTMLResponse)
def reports_page(request: Request, mode: str = "daily", db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r

    now = datetime.now()
    q = db.query(Transaction)

    if mode == "daily":
        start = datetime(now.year, now.month, now.day, 0, 0, 0)
        end = datetime(now.year, now.month, now.day, 23, 59, 59)
        title = f"Laporan Harian ({start.date()})"
    elif mode == "monthly":
        start = datetime(now.year, now.month, 1, 0, 0, 0)
        # Calculate last day of the month
        last_day = monthrange(now.year, now.month)[1]
        end = datetime(now.year, now.month, last_day, 23, 59, 59)
        title = f"Laporan Bulanan ({now.strftime('%Y-%m')})"
    else:
        start = datetime(now.year, 1, 1, 0, 0, 0)
        end = datetime(now.year, 12, 31, 23, 59, 59)
        title = f"Laporan Tahunan ({now.year})"

    trx_list = q.filter(Transaction.created_at >= start, Transaction.created_at <= end).order_by(Transaction.id.desc()).all()
    omzet = sum(t.total for t in trx_list)
    # Hitung modal (cost) dari item transaksi untuk pendapatan rill (omzet - modal penjualan)
    total_modal = 0
    for t in trx_list:
        for item in t.items:
            total_modal += (item.cost_price or 0) * (item.qty or 0)
    pendapatan_rill = omzet - total_modal

    # Pengeluaran dari upgrade stok (dari tabel stock_updates) pada periode yang sama
    stock_updates = db.query(StockUpdate).filter(StockUpdate.created_at >= start, StockUpdate.created_at <= end).all()
    pengeluaran_stok = sum(su.total_pengeluaran or 0 for su in stock_updates)

    jumlah = len(trx_list)
    
    # Format created_at untuk semua transaksi
    for trx in trx_list:
        trx.formatted_created_at = format_datetime_with_tz(trx.created_at, db, "%Y-%m-%d %H:%M")

    return templates.TemplateResponse("reports.html", {
        "request": request,
        "title": title,
        "mode": mode,
        "trx_list": trx_list,
        "omzet": format_idr(omzet),
        "pendapatan_rill": format_idr(pendapatan_rill),
        "pengeluaran_stok": format_idr(pengeluaran_stok),
        "jumlah": jumlah,
        "user": request.session["user"]
    })

@app.get("/api/reports/summary")
def get_summary_report(db: Session = Depends(get_db)):
    total_products = db.query(Product).count()
    active_products = db.query(Product).filter(Product.status == "active").count()

    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    end_of_month = datetime(now.year, now.month, monthrange(now.year, now.month)[1], 23, 59, 59)

    products_sold_this_month = (
        db.query(func.sum(TransactionItem.qty))
        .join(Transaction)
        .filter(Transaction.created_at >= start_of_month, Transaction.created_at <= end_of_month)
        .scalar()
        or 0
    )

    return {
        "total_products": total_products,
        "active_products": active_products,
        "products_sold_this_month": products_sold_this_month,
    }

@app.get("/api/reports/top_products")
def get_top_products_report(db: Session = Depends(get_db)):
    top_selling_products = (
        db.query(
            Product.name,
            func.sum(TransactionItem.qty).label("total_qty_sold")
        )
        .join(TransactionItem, Product.code == TransactionItem.product_code)
        .group_by(Product.name)
        .order_by(func.sum(TransactionItem.qty).desc())
        .limit(5)
        .all()
    )

    highest_revenue_products = (
        db.query(
            Product.name,
            func.sum(TransactionItem.subtotal).label("total_revenue")
        )
        .join(TransactionItem, Product.code == TransactionItem.product_code)
        .group_by(Product.name)
        .order_by(func.sum(TransactionItem.subtotal).desc())
        .limit(5)
        .all()
    )

    return {
        "top_selling_products": [{"name": p.name, "total_qty_sold": p.total_qty_sold} for p in top_selling_products],
        "highest_revenue_products": [{"name": p.name, "total_revenue": p.total_revenue} for p in highest_revenue_products],
    }

@app.get("/api/reports/problem_products")
def get_problem_products_report(db: Session = Depends(get_db)):
    # Produk jarang laku (misal, terjual kurang dari 5 unit sepanjang waktu)
    rarely_sold_products = (
        db.query(
            Product.name,
            func.sum(TransactionItem.qty).label("total_qty_sold")
        )
        .outerjoin(TransactionItem, Product.code == TransactionItem.product_code)
        .group_by(Product.name)
        .having(func.sum(TransactionItem.qty) < 5)
        .order_by(func.sum(TransactionItem.qty).asc())
        .limit(5)
        .all()
    )

    # Produk tidak pernah terjual
    never_sold_products = (
        db.query(Product.name)
        .outerjoin(TransactionItem, Product.code == TransactionItem.product_code)
        .filter(TransactionItem.id == None)
        .limit(5)
        .all()
    )

    return {
        "rarely_sold_products": [{"name": p.name, "total_qty_sold": p.total_qty_sold if p.total_qty_sold is not None else 0} for p in rarely_sold_products],
        "never_sold_products": [{"name": p.name} for p in never_sold_products],
    }

@app.get("/api/reports/stock")
def get_stock_report(db: Session = Depends(get_db)):
    # Produk hampir habis (misal, stok kurang dari 10)
    low_stock_products = (
        db.query(Product.name, Product.stock)
        .filter(Product.stock < 10, Product.status == "active")
        .order_by(Product.stock.asc())
        .limit(5)
        .all()
    )

    # Produk overstock (misal, stok lebih dari 100)
    overstock_products = (
        db.query(Product.name, Product.stock)
        .filter(Product.stock > 100, Product.status == "active")
        .order_by(Product.stock.desc())
        .limit(5)
        .all()
    )

    return {
        "low_stock_products": [{"name": p.name, "stock": p.stock} for p in low_stock_products],
        "overstock_products": [{"name": p.name, "stock": p.stock} for p in overstock_products],
    }

@app.get("/api/reports/sales_trend")
def get_sales_trend_report(db: Session = Depends(get_db)):
    sales_trend = (
        db.query(
            func.strftime("%Y-%m", Transaction.created_at).label("month"),
            func.sum(Transaction.total).label("total_sales")
        )
        .group_by("month")
        .order_by("month")
        .all()
    )

    return {
        "sales_trend": [{"month": s.month, "total_sales": s.total_sales} for s in sales_trend]
    }

# -------- SETTINGS ----------
@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r
    # Dapatkan pengaturan yang ada atau buat yang baru jika tidak ada
    settings = db.query(Setting).first()
    if not settings:
        settings = Setting(store_name="Nama Toko Anda", store_address="Alamat Toko Anda", store_phone="")
        db.add(settings)
        db.commit()
        db.refresh(settings)

    # Inisialisasi display_name jika belum ada
    current_user_db = db.query(User).filter(User.username == request.session["user"]["username"]).first()
    if current_user_db and current_user_db.display_name is None:
        current_user_db.display_name = current_user_db.username
        db.commit()
        db.refresh(current_user_db)
        request.session["user"]["display_name"] = current_user_db.display_name # Update session immediately

    return templates.TemplateResponse("settings.html", {"request": request, "user": request.session["user"], "settings": settings})

@app.post("/settings")
def update_settings(
    request: Request,
    store_name: str = Form(...),
    store_address: str = Form(...),
    store_phone: str = Form(...),
    timezone: str = Form(default="WIB"),
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r
    
    # Validasi timezone
    if timezone not in ["WIB", "WITA", "WIT"]:
        timezone = "WIB"
    
    settings = db.query(Setting).first()
    if not settings:
        settings = Setting()
        db.add(settings)
    settings.store_name = store_name
    settings.store_address = store_address
    settings.store_phone = store_phone
    settings.timezone = timezone
    db.commit()
    return RedirectResponse("/settings?msg=updated", status_code=302)

@app.post("/settings/update_display_name")
def update_display_name(
    request: Request,
    new_display_name: str = Form(...),
    db: Session = Depends(get_db),
):
    r = require_login(request)
    if r: return r

    user_in_db = db.query(User).filter(User.username == request.session["user"]["username"]).first()
    if user_in_db:
        user_in_db.display_name = new_display_name
        db.commit()
        db.refresh(user_in_db)
        request.session["user"]["display_name"] = user_in_db.display_name # Update session immediately
        return RedirectResponse("/settings?msg=display_name_updated", status_code=302)
    return RedirectResponse("/settings?msg=error", status_code=302)


@app.get("/settings/export_db")
def export_database(request: Request):
    """Export seluruh database SQLite (pos.db) sebagai file download."""
    r = require_login(request)
    if r: return r
    db_path = os.path.join(os.path.dirname(__file__), "pos.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="File database tidak ditemukan.")
    filename = f"pos_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return FileResponse(
        path=db_path,
        media_type="application/octet-stream",
        filename=filename,
    )


@app.post("/settings/import_db")
def import_database(
    request: Request,
    file: UploadFile = File(...),
):
    """Import database SQLite dari file upload (replace pos.db)."""
    r = require_login(request)
    if r: return r

    # Validasi ekstensi
    if not file.filename.lower().endswith((".db", ".sqlite", ".sqlite3")):
        return RedirectResponse("/settings?msg=error_import", status_code=302)

    db_dir = os.path.dirname(__file__)
    db_path = os.path.join(db_dir, "pos.db")
    tmp_path = os.path.join(db_dir, "pos_import_tmp.db")

    try:
        # Simpan upload ke file sementara
        with open(tmp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Tutup semua koneksi aktif ke SQLite sebelum replace
        engine.dispose()

        # Replace database lama dengan yang baru
        shutil.move(tmp_path, db_path)

        return RedirectResponse("/settings?msg=imported", status_code=302)
    except Exception as e:
        # Bersihkan tmp jika ada
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except:
            pass
        print(f"Error import database: {e}")
        return RedirectResponse("/settings?msg=error_import", status_code=302)

@app.post("/settings/clear_database")
def clear_database(request: Request, db: Session = Depends(get_db)):
    """Hapus semua data dari database (produk, transaksi, stock updates, dll)"""
    r = require_login(request)
    if r: return r
    
    try:
        # Hapus semua data dalam urutan yang benar (menghindari foreign key constraint)
        db.query(TransactionItem).delete()
        db.query(StockUpdate).delete()
        db.query(Transaction).delete()
        db.query(Product).delete()
        # Settings dan User tidak dihapus untuk keamanan
        db.commit()
        return RedirectResponse("/settings?msg=database_cleared", status_code=302)
    except Exception as e:
        db.rollback()
        return RedirectResponse("/settings?msg=error_clear", status_code=302)

def _apply_excel_styling(workbook, writer, transactions, total_pendapatan, total_pengeluaran, db):
    """Fungsi helper untuk apply styling ke semua sheet Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Biru
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Roboto")  # Font lebih kecil
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Fungsi helper untuk styling sheet dengan desain yang konsisten
    def style_sheet_consistent(ws, sheet_name):
        """Apply styling konsisten ke semua sheet"""
        if ws.max_row <= 1:
            return  # Skip jika tidak ada data
        
        ws.freeze_panes = 'A2'
        
        # Style header - semua sheet menggunakan warna biru yang sama
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        
        # Set row height untuk header
        ws.row_dimensions[1].height = 20
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border_style
                
                # Format angka - cek semua kolom yang mungkin berisi angka
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                elif cell.column_letter == 'A':  # Kolom pertama biasanya teks
                    cell.alignment = left_align
                else:
                    cell.alignment = left_align
        
        # Auto-width untuk semua kolom
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_str = str(cell.value)
                        if len(cell_str) > max_length:
                            max_length = len(cell_str)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            if adjusted_width > 0:
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply styling konsisten ke semua sheet
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        style_sheet_consistent(ws, sheet_name)

@app.get("/api/reports/export_excel")
def export_reports_excel(request: Request, mode: str = "daily", db: Session = Depends(get_db)):
    r = require_login(request)
    if r: return r

    now = datetime.now()
    start, end, title_prefix = None, None, ""

    if mode == "daily":
        start = datetime(now.year, now.month, now.day, 0, 0, 0)
        end = datetime(now.year, now.month, now.day, 23, 59, 59)
        title_prefix = f"Harian-{start.date()}"
    elif mode == "monthly":
        start = datetime(now.year, now.month, 1, 0, 0, 0)
        last_day = monthrange(now.year, now.month)[1]
        end = datetime(now.year, now.month, last_day, 23, 59, 59)
        title_prefix = f"Bulanan-{now.strftime('%Y-%m')}"
    else:
        start = datetime(now.year, 1, 1, 0, 0, 0)
        end = datetime(now.year, 12, 31, 23, 59, 59)
        title_prefix = f"Tahunan-{now.year}"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        # Ringkasan Transaksi - tetap gunakan tanggal lengkap untuk mengetahui tgl pembelian
        transactions = db.query(Transaction).filter(Transaction.created_at >= start, Transaction.created_at <= end).all()
        
        df_trx = pd.DataFrame([{
            "TRX_NO": t.trx_no,
            "TANGGAL": t.created_at.strftime("%Y-%m-%d %H:%M"),
            "BULAN_TAHUN": t.created_at.strftime("%Y-%m"),  # Tambah kolom bulan-tahun untuk ringkasan
            "KASIR": t.cashier,
            "METODE_BAYAR": t.payment_method,
            "TOTAL": t.total,
            "BAYAR": t.paid,
            "KEMBALI": t.change,
        } for t in transactions])
        df_trx.to_excel(writer, sheet_name='Transaksi', index=False)
        
        # Hitung total pendapatan dan total pengeluaran
        total_pendapatan = sum(t.total for t in transactions)
        total_pengeluaran = 0
        for t in transactions:
            for item in t.items:
                total_pengeluaran += item.cost_price * item.qty
        
        # Tambahkan total pengeluaran dari stock updates dalam periode yang sama
        stock_updates = db.query(StockUpdate).filter(
            StockUpdate.created_at >= start,
            StockUpdate.created_at <= end
        ).all()
        total_pengeluaran_stock = sum(su.total_pengeluaran for su in stock_updates)
        total_pengeluaran += total_pengeluaran_stock
        
        # Tambah sheet Ringkasan Keuangan
        df_keuangan = pd.DataFrame([{
            "Keterangan": "Total Pendapatan",
            "Nilai": total_pendapatan
        }, {
            "Keterangan": "Total Pengeluaran (Penjualan)",
            "Nilai": total_pengeluaran - total_pengeluaran_stock
        }, {
            "Keterangan": "Total Pengeluaran (Update Stok)",
            "Nilai": total_pengeluaran_stock
        }, {
            "Keterangan": "Total Pengeluaran",
            "Nilai": total_pengeluaran
        }, {
            "Keterangan": "Laba/Rugi",
            "Nilai": total_pendapatan - total_pengeluaran
        }])
        df_keuangan.to_excel(writer, sheet_name='Ringkasan Keuangan', index=False)
        
        # Tambah sheet Detail Update Stok jika ada
        if stock_updates:
            df_stock_updates = pd.DataFrame([{
                "TANGGAL": su.created_at.strftime("%Y-%m-%d %H:%M"),
                "KODE_PRODUK": su.product_code,
                "NAMA_PRODUK": su.product_name,
                "STOK_LAMA": su.old_stock,
                "STOK_BARU": su.new_stock,
                "STOK_DITAMBAH": su.stock_added,
                "HARGA_ASLI": su.cost_price,
                "TOTAL_PENGELUARAN": su.total_pengeluaran,
                "DIPERBARUI_OLEH": su.updated_by or "System"
            } for su in stock_updates])
            df_stock_updates.to_excel(writer, sheet_name='Detail Update Stok', index=False)

        # Ringkasan Produk (efisien - hanya data penting)
        summary_data = get_summary_report(db)
        df_summary = pd.DataFrame([{
            "Ringkasan": "Total Produk", "Nilai": summary_data["total_products"]},
            {"Ringkasan": "Produk Terjual Bulan Ini", "Nilai": summary_data["products_sold_this_month"]},
        ])
        df_summary.to_excel(writer, sheet_name='Ringkasan Produk', index=False)

        # Top Produk Paling Laku
        top_products = get_top_products_report(db)
        df_top_selling = pd.DataFrame(top_products["top_selling_products"])
        df_top_selling.rename(columns={'name': 'Nama Produk', 'total_qty_sold': 'Total Terjual'})
        df_top_selling.to_excel(writer, sheet_name='Top Produk Laku', index=False)

        # Top Produk Omzet Tertinggi
        df_highest_revenue = pd.DataFrame(top_products["highest_revenue_products"])
        df_highest_revenue.rename(columns={'name': 'Nama Produk', 'total_revenue': 'Total Omzet'})
        df_highest_revenue.to_excel(writer, sheet_name='Top Produk Omzet', index=False)

        # Produk Bermasalah
        problem_products = get_problem_products_report(db)
        df_rarely_sold = pd.DataFrame(problem_products["rarely_sold_products"])
        df_rarely_sold.rename(columns={'name': 'Nama Produk', 'total_qty_sold': 'Jumlah Terjual'})
        df_rarely_sold.to_excel(writer, sheet_name='Jarang Laku', index=False)

        df_never_sold = pd.DataFrame(problem_products["never_sold_products"])
        df_never_sold.rename(columns={'name': 'Nama Produk'})
        df_never_sold.to_excel(writer, sheet_name='Tidak Terjual', index=False)

        # Stok Produk
        stock_report = get_stock_report(db)
        df_low_stock = pd.DataFrame(stock_report["low_stock_products"])
        df_low_stock.rename(columns={'name': 'Nama Produk', 'stock': 'Stok'})
        df_low_stock.to_excel(writer, sheet_name='Stok Hampir Habis', index=False)

        df_overstock = pd.DataFrame(stock_report["overstock_products"])
        df_overstock.rename(columns={'name': 'Nama Produk', 'stock': 'Stok'})
        df_overstock.to_excel(writer, sheet_name='Overstock', index=False)

        # Tren Penjualan
        sales_trend = get_sales_trend_report(db)
        df_sales_trend = pd.DataFrame(sales_trend["sales_trend"])
        df_sales_trend.rename(columns={'month': 'Bulan', 'total_sales': 'Total Penjualan'})
        df_sales_trend.to_excel(writer, sheet_name='Tren Penjualan', index=False)
        
        # Apply styling ke semua sheet
        _apply_excel_styling(workbook, writer, transactions, total_pendapatan, total_pengeluaran, db)

    output.seek(0)
    file_name = f"Laporan_Penjualan_{title_prefix}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename={file_name}"
    })
